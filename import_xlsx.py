#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Import completo da NUOVO CRM ECOTRENTINO (1).xlsx"""
import sys, os, re
from datetime import datetime, date

os.chdir(r'C:\Users\Alessio Osele\ecotrentino-crm')
sys.path.insert(0, r'C:\Users\Alessio Osele\ecotrentino-crm')

import openpyxl
from sqlalchemy import text

from database import SessionLocal
from models import Azienda, Contatto, Opportunita, Attivita

XLSX = r'C:\Users\Alessio Osele\Desktop\NUOVO CRM ECOTRENTINO (1).xlsx'

# ── MAPPINGS ──────────────────────────────────────────────────────────────────

SETTORE_MAP = {
    # Molitura
    "industria molitoria": "Molitura",
    "molitura": "Molitura",
    "macchinari per la molitura": "Molitura",
    "macchinari per  molitura, essicazione": "Molitura",
    "macchinari per molitura, essicazione": "Molitura",
    "macinazione e molitura": "Molitura",
    "macinazione": "Molitura",
    "macinzione e molitura": "Molitura",
    "filtrazionee molitira": "Molitura",
    # Pharma
    "chimico-farmaceutico": "Pharma",
    "macchinari per industria farmaceutica": "Pharma",
    "pharma": "Pharma",
    # Meccanica / Engineering
    "meccanica": "Meccanica",
    "macchinari per l'industria": "Meccanica",
    "impianti industriali": "Meccanica",
    "impianti induistriali": "Meccanica",
    "impianti industraili": "Meccanica",
    "impianti induistriali ": "Meccanica",
    "impiantista": "Meccanica",
    "engineering": "Engineering",
    "engineering e produzione": "Engineering",
    "edilizia industriale e montaggio": "Meccanica",
    "montaggi industriali": "Meccanica",
    "montaggio industraiale": "Meccanica",
    # Metalli
    "metalli": "Metalli",
    "siderurgico": "Metalli",
    # Recycling
    "recycling": "Recycling",
    # Alimentare
    "alimentare": "Alimentare",
    # Legno
    "legno": "Legno",
    # Aeraulica / Ventilazione
    "aeraulica": "Aeraulica",
    "ventilazione": "Aeraulica",
    # Filtrazione
    "filtrazione": "Filtrazione",
    "filtrazione e manutenzione": "Filtrazione",
    "purificazione": "Filtrazione",
    # Chimica
    "chimica": "Chimica",
    # Tubazioni
    "tuberia": "Tubazioni",
    "tubazioni": "Tubazioni",
    "vendita tubi": "Tubazioni",
    # Verniciatura
    "verniciatura": "Verniciatura",
    "sabbiatura": "Verniciatura",
    # Biomassa
    "biomassa": "Biomassa",
    # Canalista
    "canalista": "Canalista",
    # Edilizia
    "cemento": "Edilizia",
    "cantieristica": "Edilizia",
    # Manutenzione
    "manutenzione": "Manutenzione",
    # Sicurezza
    "sicurezza infrastrutture": "Sicurezza",
    # Altro
    "altro": "Altro",
}

STATO_MAP = {
    "lead": "freddo",
    "prima telefonata": "contattato",
    "primo contatto": "contattato",
    "in trattativa": "trattativa",
    "cliente acquisito": "cliente",
}

PROV_REGIONE = {
    "AG": "Sicilia", "AL": "Piemonte", "AN": "Marche", "AO": "Valle d'Aosta",
    "AP": "Marche", "AQ": "Abruzzo", "AR": "Toscana", "AT": "Piemonte",
    "AV": "Campania", "BA": "Puglia", "BG": "Lombardia", "BI": "Piemonte",
    "BL": "Veneto", "BN": "Campania", "BO": "Emilia-Romagna", "BR": "Puglia",
    "BS": "Lombardia", "BT": "Puglia", "BZ": "Trentino-Alto Adige",
    "CA": "Sardegna", "CB": "Molise", "CE": "Campania", "CH": "Abruzzo",
    "CL": "Sicilia", "CN": "Piemonte", "CO": "Lombardia", "CR": "Lombardia",
    "CS": "Calabria", "CT": "Sicilia", "CZ": "Calabria", "EN": "Sicilia",
    "FC": "Emilia-Romagna", "FE": "Emilia-Romagna", "FG": "Puglia",
    "FI": "Toscana", "FM": "Marche", "FR": "Lazio", "GE": "Liguria",
    "GO": "Friuli-Venezia Giulia", "GR": "Toscana", "IM": "Liguria",
    "IS": "Molise", "KR": "Calabria", "LC": "Lombardia", "LE": "Puglia",
    "LI": "Toscana", "LO": "Lombardia", "LT": "Lazio", "LU": "Toscana",
    "MB": "Lombardia", "MC": "Marche", "ME": "Sicilia", "MI": "Lombardia",
    "MN": "Lombardia", "MO": "Emilia-Romagna", "MS": "Toscana", "MT": "Basilicata",
    "NA": "Campania", "NO": "Piemonte", "NU": "Sardegna", "OR": "Sardegna",
    "PA": "Sicilia", "PC": "Emilia-Romagna", "PD": "Veneto", "PE": "Abruzzo",
    "PG": "Umbria", "PI": "Toscana", "PN": "Friuli-Venezia Giulia",
    "PO": "Toscana", "PR": "Emilia-Romagna", "PT": "Toscana", "PU": "Marche",
    "PV": "Lombardia", "PZ": "Basilicata", "RA": "Emilia-Romagna",
    "RC": "Calabria", "RE": "Emilia-Romagna", "RG": "Sicilia", "RI": "Lazio",
    "RM": "Lazio", "RN": "Emilia-Romagna", "RO": "Veneto", "SA": "Campania",
    "SI": "Toscana", "SO": "Lombardia", "SP": "Liguria", "SR": "Sicilia",
    "SS": "Sardegna", "SU": "Sardegna", "SV": "Liguria", "TA": "Puglia",
    "TE": "Abruzzo", "TN": "Trentino-Alto Adige", "TO": "Piemonte",
    "TP": "Sicilia", "TR": "Umbria", "TS": "Friuli-Venezia Giulia",
    "TV": "Veneto", "UD": "Friuli-Venezia Giulia", "VA": "Lombardia",
    "VB": "Piemonte", "VC": "Piemonte", "VE": "Veneto", "VI": "Veneto",
    "VR": "Veneto", "VT": "Lazio", "VV": "Calabria",
}

STRATEGIC_KEYWORDS = [
    "OEM", "Da approcciare", "Da monitorare", "Partnership di filiera",
    "MECSPE", "Fornitura tecnica /", "Contatto OEM", "Contatto secondario",
    "Contatto tecnico",
]

# ── HELPERS ───────────────────────────────────────────────────────────────────

def normalize_settore(val):
    if not val:
        return "Altro"
    return SETTORE_MAP.get(val.strip().lower(), val.strip())

def map_stato(val):
    if not val:
        return "freddo"
    return STATO_MAP.get(val.strip().lower(), "freddo")

def parse_citta_provincia(val):
    if not val:
        return None, None
    val = str(val).strip()
    m = re.search(r'\(([A-Z]{2,3})\)', val)
    if m:
        prov = m.group(1)[:2]
        citta = val[:m.start()].strip().rstrip(',').strip()
        return citta, prov
    return val, None

def parse_indirizzo(val):
    if not val:
        return None
    val = str(val).strip()
    val = re.sub(r'\s*[-–]\s*\d{4,5}\s*$', '', val).strip()
    return val or None

def is_strategic_note(val):
    if not val:
        return False
    s = str(val)
    if len(s) > 100:
        return True
    return any(kw in s for kw in STRATEGIC_KEYWORDS)

def parse_prodotto(val):
    """Returns (prodotto_interesse, note_extra)"""
    if not val:
        return None, None
    s = str(val).strip()
    if not s:
        return None, None
    if is_strategic_note(s):
        return None, s
    return s[:255], None

def parse_contatto_name(val):
    """'Roberto Resi - Acquisti' → (nome, cognome, ruolo)"""
    if not val:
        return None, None, None
    val = str(val).strip()
    ruolo = None
    if ' - ' in val:
        parts = val.split(' - ', 1)
        val = parts[0].strip()
        ruolo = parts[1].strip()
    role_kws = ['Titolare', 'Direttore', 'Tecnico', 'Commerciale', 'Acquisti',
                'Responsabile', 'Manager', 'Amministratore', 'CEO', 'CTO']
    words = val.split()
    if len(words) > 2 and words[-1] in role_kws:
        if not ruolo:
            ruolo = words[-1]
        words = words[:-1]
    if not words:
        return None, None, None
    if len(words) == 1:
        return words[0], '-', ruolo
    return ' '.join(words[:-1]), words[-1], ruolo

def parse_contatto_diretto(val):
    """Returns (email, telefono)"""
    if not val:
        return None, None
    s = str(val).strip().replace('@@', '@')
    if '@' in s:
        return s, None
    return None, s[:20]

def to_date(val):
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    return None

def to_datetime(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    return None

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("=== IMPORT CRM ECOTRENTINO ===\n")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    db = SessionLocal()

    # 1. SVUOTA
    print("Svuoto il database...")
    db.execute(text("DELETE FROM attivita"))
    db.execute(text("DELETE FROM opportunita"))
    db.execute(text("DELETE FROM contatti"))
    db.execute(text("DELETE FROM aziende"))
    db.commit()
    print("Database svuotato.\n")

    # 2. RACCOGLI RECORD UNICI
    records = []
    seen = set()

    ws_db = wb['Database']
    for r in list(ws_db.iter_rows(values_only=True))[1:]:
        if r[0] and str(r[0]).strip():
            key = str(r[0]).strip().lower()
            if key not in seen:
                seen.add(key)
                records.append((r, None))

    for sheet, label in [('Molitura', 'Molitura'), ('Pharma', 'Pharma'),
                          ('Edilizia', 'Edilizia'), ('Militare', 'Militare')]:
        ws = wb[sheet]
        for r in list(ws.iter_rows(values_only=True))[1:]:
            if r[0] and str(r[0]).strip():
                key = str(r[0]).strip().lower()
                if key not in seen:
                    seen.add(key)
                    records.append((r, label))

    print(f"Record totali da importare: {len(records)}\n")

    # 3. IMPORT AZIENDE / CONTATTI / OPPORTUNITA
    azienda_map = {}
    warns = []

    def lookup_azienda(nome):
        key = nome.strip().lower()
        if key in azienda_map:
            return azienda_map[key]
        # prova match parziale: cerca chiavi che iniziano con 'key'
        for k, v in azienda_map.items():
            if k.startswith(key) or key.startswith(k):
                return v
        return None

    for i, (r, etichetta) in enumerate(records, 1):
        ragione_sociale = str(r[0]).strip()
        attivita_desc   = str(r[1]).strip()[:255] if r[1] else None
        indirizzo       = parse_indirizzo(r[2])
        citta, prov     = parse_citta_provincia(r[3])
        regione         = PROV_REGIONE.get(prov) if prov else None
        website         = str(r[4]).strip() if r[4] else None
        email_az        = str(r[5]).strip() if r[5] else None
        tel_az          = str(r[6]).strip()[:20] if r[6] else None
        contatto_dir    = r[7]
        contatto_raw    = r[8]
        settore_raw     = r[9]
        prodotto_raw    = r[10]
        fonte_lead      = str(r[11]).strip() if r[11] else None
        stato_raw       = str(r[12]).strip() if r[12] else None
        ultimo_cont     = r[13]
        prossimo_fu     = r[14]
        feedback        = str(r[15]).strip() if r[15] else None
        ordine_raw      = r[16]
        commessa_raw    = r[17]
        offerte_coll    = r[20]

        tipo            = normalize_settore(settore_raw)
        prodotto, note_strat = parse_prodotto(prodotto_raw)
        ordine          = str(ordine_raw).strip().lower() in ('sì', 'si', 'yes', '1', 'true') if ordine_raw else False
        commessa_euro   = float(commessa_raw) if commessa_raw and commessa_raw != 0 else None

        az = Azienda(
            ragione_sociale    = ragione_sociale,
            attivita_descrizione = attivita_desc,
            indirizzo          = indirizzo,
            citta              = citta,
            provincia          = prov,
            regione            = regione,
            email_aziendale    = email_az,
            telefono_aziendale = tel_az,
            website            = website,
            tipo               = tipo,
            prodotto_interesse = prodotto,
            fonte_lead         = fonte_lead,
            ordine             = ordine,
            commessa_euro      = commessa_euro,
            etichetta          = etichetta,
        )
        db.add(az)
        db.flush()
        azienda_map[ragione_sociale.lower()] = az.id

        # CONTATTO
        email_c, tel_c = parse_contatto_diretto(contatto_dir)
        nome, cognome, ruolo = parse_contatto_name(contatto_raw)
        cont_id = None
        if nome and cognome:
            cont = Contatto(
                id_azienda = az.id,
                nome       = nome,
                cognome    = cognome,
                ruolo      = ruolo,
                email      = email_c,
                telefono   = tel_c,
            )
            db.add(cont)
            db.flush()
            cont_id = cont.id

        # OPPORTUNITA
        stato     = map_stato(stato_raw)
        note_opp  = feedback or ''
        if note_strat:
            note_opp = (note_opp + '\n' + note_strat).strip() if note_opp else note_strat
        titolo    = f"{prodotto} — {ragione_sociale}" if prodotto else f"Pipeline {ragione_sociale}"
        off_str   = str(int(offerte_coll)) if offerte_coll and offerte_coll != 0 else None

        opp = Opportunita(
            id_azienda           = az.id,
            id_contatto          = cont_id,
            titolo               = titolo[:255],
            stato                = stato,
            data_ultimo_contatto = to_datetime(ultimo_cont),
            prossimo_followup    = to_date(prossimo_fu),
            offerte_collegate    = off_str,
            note                 = note_opp or None,
        )
        db.add(opp)

        if i % 50 == 0:
            db.commit()
            print(f"  {i}/{len(records)} importati...")

    db.commit()
    print(f"  {len(records)}/{len(records)} importati.")

    # 4. ATTIVITA
    print("\nImport attività...")
    ws_att  = wb['Attività']
    rows_att = list(ws_att.iter_rows(values_only=True))
    att_count = 0
    for r in rows_att[2:]:
        if not r[0] or not r[3]:
            continue
        cliente   = str(r[0]).strip()
        azione    = str(r[1]).strip() if r[1] else None
        data_att  = r[3]
        stato_att = str(r[4]).strip() if r[4] else None
        note_att  = str(r[6]).strip() if r[6] else None
        az_id     = lookup_azienda(cliente)
        if not az_id:
            warns.append(f"Attivita: azienda '{cliente}' non trovata")
            continue
        tipo_att = "offerta" if azione and "ordine" in azione.lower() else "email"
        att = Attivita(
            id_azienda  = az_id,
            tipo        = tipo_att,
            data        = to_datetime(data_att) or datetime.now(),
            descrizione = azione,
            esito       = stato_att,
        )
        db.add(att)
        att_count += 1
    db.commit()
    print(f"  Attività importate: {att_count}")

    # 5. OFFERTE come note
    print("\nImport offerte come note...")
    ws_off   = wb['Offerte']
    rows_off = list(ws_off.iter_rows(values_only=True))
    off_count = 0
    for r in rows_off[1:]:
        if not r[0] or not r[1]:
            continue
        id_off    = str(r[0]).strip()
        cliente   = str(r[1]).strip()
        oggetto   = str(r[3]).strip() if r[3] else '-'
        importo   = r[4]
        stato_off = str(r[5]).strip() if r[5] else '-'
        note_off  = str(r[7]).strip() if r[7] else ''
        az_id     = lookup_azienda(cliente)
        if not az_id:
            warns.append(f"Offerta: azienda '{cliente}' non trovata")
            continue
        az = db.query(Azienda).filter(Azienda.id == az_id).first()
        nota = f"[{id_off}] {oggetto} | Stato: {stato_off}"
        if importo:
            nota += f" | €{importo}"
        if note_off:
            nota += f" | {note_off}"
        az.note = ((az.note or '') + '\n' + nota).strip()
        off_count += 1
    db.commit()
    print(f"  Offerte inserite come note: {off_count}")

    # STATS
    total_az   = db.query(Azienda).count()
    total_cont = db.query(Contatto).count()
    total_opp  = db.query(Opportunita).count()
    total_att  = db.query(Attivita).count()

    print(f"""
=== RISULTATO FINALE ===
  Aziende:      {total_az}
  Contatti:     {total_cont}
  Opportunità:  {total_opp}
  Attività:     {total_att}
""")

    if warns:
        print("AVVISI:")
        for w in warns:
            print(f"  [!] {w}")

    db.close()
    print("Import completato con successo!")

if __name__ == '__main__':
    main()
